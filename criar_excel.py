"""
Gera o arquivo Imile Delivery_Operacoes.xlsx para o teste de Analista Sênior.
4 abas: Entregas (5000 linhas), Clientes, Parametros, Motoristas
"""
import random
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

random.seed(42)

# ── Dimensões ─────────────────────────────────────────────────
CDS = [
    "SP-Centro", "SP-Zona Sul", "SP-ABC", "SP-Leste",
    "RJ-Centro", "RJ-Zona Norte", "BH-Centro", "POA-Centro",
]
SEGMENTOS = ["B2B", "B2C", "E-commerce", "Farmácias"]
REGIOES = {
    "SP-Centro": "Sudeste",  "SP-Zona Sul": "Sudeste",
    "SP-ABC":    "Sudeste",  "SP-Leste":    "Sudeste",
    "RJ-Centro": "Sudeste",  "RJ-Zona Norte": "Sudeste",
    "BH-Centro": "Sudeste",  "POA-Centro":  "Sul",
}
VEICULOS = ["Moto", "Carro", "Van"]
SLA_MAP  = {"B2B": 2, "B2C": 3, "E-commerce": 2, "Farmácias": 1}

NOMES_MOTORISTAS = [
    "Carlos Silva", "Ana Souza", "Pedro Lima", "Julia Costa",
    "Marcos Oliveira", "Fernanda Santos", "Rafael Pereira", "Camila Rocha",
    "Bruno Ferreira", "Leticia Alves", "Diego Martins", "Patrícia Gomes",
    "Ricardo Nunes", "Vanessa Cruz", "Thiago Barbosa", "Natália Freitas",
    "Felipe Cardoso", "Aline Ribeiro", "Eduardo Monteiro", "Simone Castro",
    "Gabriel Teixeira", "Larissa Moreira", "Henrique Dias", "Priscila Araujo",
    "Lucas Mendes", "Daniela Lopes", "André Cavalcanti", "Sabrina Pinto",
    "Igor Nascimento", "Tatiana Carvalho", "Rodrigo Vasconcelos", "Mariana Faria",
    "Guilherme Azevedo", "Isabela Correia", "Renato Borges", "Amanda Melo",
    "Vinícius Ramos", "Juliana Cunha", "Leandro Machado", "Cristina Torres",
]

NOMES_CLIENTES = [
    "Magazine Luiza", "Americanas", "Shoptime", "Casas Bahia",
    "Drogasil", "Farmácias Pacheco", "iFood", "Rappi",
    "Lojas Renner", "C&A", "NetFarma", "Raia Drogasil",
    "Grupo Pão de Açúcar", "Carrefour", "Extra", "Assaí",
    "Natura", "Boticário", "Avon", "Mary Kay",
]

def rand_date(start: date, end: date) -> date:
    return start + timedelta(days=random.randint(0, (end - start).days))

# ── Motoristas ───────────────────────────────────────────────
motoristas = []
for i, nome in enumerate(NOMES_MOTORISTAS, 1):
    cd = random.choice(CDS)
    veiculo = random.choices(VEICULOS, weights=[5, 3, 2])[0]
    motoristas.append({"id": i, "nome": nome, "cd": cd, "veiculo": veiculo})

# ── Clientes ─────────────────────────────────────────────────
clientes = []
for i, nome in enumerate(NOMES_CLIENTES, 1):
    seg = random.choice(SEGMENTOS)
    # Intencionalmente com problemas de espaço/case para o exercício BI Q1
    regiao_raw = random.choice(["Sudeste", "  sudeste ", "SUDESTE", "Sul", " sul", "Nordeste", "nordeste  "])
    clientes.append({"id": i, "nome": nome, "segmento": seg, "regiao": regiao_raw})

# ── Entregas ─────────────────────────────────────────────────
inicio = date(2024, 1, 2)
fim_coleta = date(2024, 3, 31)
entregas = []
for i in range(1, 5001):
    motor = random.choice(motoristas)
    cli   = random.choice(clientes)
    seg   = cli["segmento"]
    sla   = SLA_MAP[seg]
    cd    = motor["cd"]

    data_coleta   = rand_date(inicio, fim_coleta)
    data_prevista = data_coleta + timedelta(days=sla)

    rnd = random.random()
    if rnd < 0.03:                     # 3%  Não entregue
        data_entrega = None
    elif rnd < 0.13:                   # 10% Atraso crítico (>2 dias)
        dias_atraso = random.randint(3, 10)
        data_entrega = data_prevista + timedelta(days=dias_atraso)
    elif rnd < 0.23:                   # 10% Atraso leve (1-2 dias)
        dias_atraso = random.randint(1, 2)
        data_entrega = data_prevista + timedelta(days=dias_atraso)
    else:                              # 77% No prazo
        days_before = random.randint(0, sla)
        data_entrega = data_coleta + timedelta(days=days_before)

    km = round(random.uniform(2, 45), 1)
    custo_oc = round(random.uniform(0, 120), 2)

    entregas.append({
        "id":              i,
        "data_coleta":     data_coleta,
        "data_prevista":   data_prevista,
        "data_entrega":    data_entrega,
        "cliente_id":      cli["id"],
        "segmento":        seg,
        "CD":              cd,
        "motorista_id":    motor["id"],
        "tipo_veiculo":    motor["veiculo"],
        "km_percorrido":   km,
        "custo_ocorrencia": custo_oc,
    })

# ── Estilo helpers ────────────────────────────────────────────
PURPLE = "534AB7"
PURPLE_L = "EEEDFE"
TEAL   = "1D9E75"
TEAL_L = "E1F5EE"
AMBER  = "BA7517"
AMBER_L = "FAEEDA"
GRAY_L = "F1EFE8"

def hdr(ws, row, col, value, color=PURPLE, fg="FFFFFF", bold=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=fg, size=11)
    c.fill = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return c

def cell_style(ws, row, col, value, fill=None, number_format=None, align="left"):
    c = ws.cell(row=row, column=col, value=value)
    thin = Side(style="thin", color="DDDDDD")
    c.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    c.alignment = Alignment(horizontal=align, vertical="center")
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if number_format:
        c.number_format = number_format
    return c

# ── Workbook ──────────────────────────────────────────────────
wb = openpyxl.Workbook()

# ═══ ABA ENTREGAS ════════════════════════════════════════════
ws = wb.active
ws.title = "Entregas"

COLS_ENT = [
    "id_entrega", "data_coleta", "data_prevista", "data_entrega",
    "cliente_id", "segmento", "CD", "motorista_id", "tipo_veiculo",
    "km_percorrido", "custo_ocorrencia",
]
col_colors = [PURPLE]*len(COLS_ENT)
for ci, col in enumerate(COLS_ENT, 1):
    hdr(ws, 1, ci, col)

date_fmt = "DD/MM/YYYY"
for ri, e in enumerate(entregas, 2):
    row_fill = GRAY_L if ri % 2 == 0 else "FFFFFF"
    cell_style(ws, ri, 1,  e["id"],             fill=row_fill, align="center")
    cell_style(ws, ri, 2,  e["data_coleta"],     fill=row_fill, number_format=date_fmt, align="center")
    cell_style(ws, ri, 3,  e["data_prevista"],   fill=row_fill, number_format=date_fmt, align="center")
    cell_style(ws, ri, 4,  e["data_entrega"],    fill=row_fill, number_format=date_fmt, align="center")
    cell_style(ws, ri, 5,  e["cliente_id"],      fill=row_fill, align="center")
    cell_style(ws, ri, 6,  e["segmento"],        fill=row_fill)
    cell_style(ws, ri, 7,  e["CD"],              fill=row_fill)
    cell_style(ws, ri, 8,  e["motorista_id"],    fill=row_fill, align="center")
    cell_style(ws, ri, 9,  e["tipo_veiculo"],    fill=row_fill)
    cell_style(ws, ri, 10, e["km_percorrido"],   fill=row_fill, number_format="0.0", align="right")
    cell_style(ws, ri, 11, e["custo_ocorrencia"],fill=row_fill, number_format="R$ #,##0.00", align="right")

# Larguras
col_widths = [10, 14, 14, 14, 12, 14, 14, 14, 14, 14, 16]
for ci, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w
ws.row_dimensions[1].height = 30
ws.freeze_panes = "A2"

# ═══ ABA CLIENTES ════════════════════════════════════════════
wc = wb.create_sheet("Clientes")
COLS_CLI = ["cliente_id", "nome_cliente", "segmento", "regiao"]
for ci, col in enumerate(COLS_CLI, 1):
    hdr(wc, 1, ci, col, color=TEAL)

for ri, cl in enumerate(clientes, 2):
    row_fill = TEAL_L if ri % 2 == 0 else "FFFFFF"
    cell_style(wc, ri, 1, cl["id"],       fill=row_fill, align="center")
    cell_style(wc, ri, 2, cl["nome"],     fill=row_fill)
    cell_style(wc, ri, 3, cl["segmento"], fill=row_fill)
    cell_style(wc, ri, 4, cl["regiao"],   fill=row_fill)

for ci, w in enumerate([12, 30, 15, 20], 1):
    wc.column_dimensions[get_column_letter(ci)].width = w
wc.row_dimensions[1].height = 28
wc.freeze_panes = "A2"

# ═══ ABA PARAMETROS ══════════════════════════════════════════
wp = wb.create_sheet("Parametros")

# Seção 1 — SLA por segmento (colunas A–B)
hdr(wp, 1, 1, "Segmento",  color=AMBER)
hdr(wp, 1, 2, "SLA_Dias",  color=AMBER)
sla_data = [("B2B", 2), ("B2C", 3), ("E-commerce", 2), ("Farmácias", 1)]
for ri, (seg, dias) in enumerate(sla_data, 2):
    cell_style(wp, ri, 1, seg,  fill=AMBER_L)
    cell_style(wp, ri, 2, dias, fill=AMBER_L, align="center")

# Seção 2 — Custo por km por tipo de veículo (colunas D–E, pulando C)
hdr(wp, 1, 4, "tipo_veiculo", color=PURPLE)
hdr(wp, 1, 5, "custo_km",     color=PURPLE)
custo_km_data = [("Moto", 1.20), ("Carro", 1.85), ("Van", 2.50)]
for ri, (v, custo) in enumerate(custo_km_data, 2):
    cell_style(wp, ri, 4, v,     fill=PURPLE_L)
    cell_style(wp, ri, 5, custo, fill=PURPLE_L, number_format="R$ #,##0.00", align="right")

# Larguras
for ci, w in enumerate([16, 12, 6, 16, 12], 1):
    wp.column_dimensions[get_column_letter(ci)].width = w
wp.row_dimensions[1].height = 28

# Nota orientadora
wp.cell(row=7, column=1, value="* SLA_Dias = prazo máximo em dias corridos por segmento")
wp.cell(row=7, column=1).font = Font(italic=True, color="888888", size=9)
wp.cell(row=8, column=4, value="* custo_km = R$/km percorrido por tipo de veículo")
wp.cell(row=8, column=4).font = Font(italic=True, color="888888", size=9)

# ═══ ABA MOTORISTAS ══════════════════════════════════════════
wm = wb.create_sheet("Motoristas")
COLS_MOT = ["motorista_id", "nome", "CD", "tipo_veiculo"]
for ci, col in enumerate(COLS_MOT, 1):
    hdr(wm, 1, ci, col, color=PURPLE)

for ri, m in enumerate(motoristas, 2):
    row_fill = PURPLE_L if ri % 2 == 0 else "FFFFFF"
    cell_style(wm, ri, 1, m["id"],      fill=row_fill, align="center")
    cell_style(wm, ri, 2, m["nome"],    fill=row_fill)
    cell_style(wm, ri, 3, m["cd"],      fill=row_fill)
    cell_style(wm, ri, 4, m["veiculo"], fill=row_fill)

for ci, w in enumerate([14, 26, 16, 14], 1):
    wm.column_dimensions[get_column_letter(ci)].width = w
wm.row_dimensions[1].height = 28
wm.freeze_panes = "A2"

# ── Salvar ────────────────────────────────────────────────────
out = r"C:\Users\IMILE-TI\Desktop\Teste-Anl-Sr\Imile Delivery_Operacoes.xlsx"
wb.save(out)
print(f"Arquivo criado: {out}")
print(f"  Entregas:   {len(entregas)} linhas")
print(f"  Clientes:   {len(clientes)} linhas")
print(f"  Motoristas: {len(motoristas)} linhas")
print(f"  Parametros: SLA (4 seg) + custo_km (3 veículos)")
